import csv 
import openpyxl
import base64
import json,os


"""
------------------基本函数---------------------
1. load_words_json(rule_excel_path,sheet_num) 从excel加载关键词，获取关键词id
2. generate_complete_expr(expr) 将字典类型表达式转为字符串类型
3. trans_sensitive_word_csv(rule_excel_path,sheet_num) 将excel中的敏感词保存至csv
4. trans_data_rule_csv(rule_excel_path,sheet_num) 将excel中的分级分类规则保存至csv
5. generate_test_json 生成用于测试的json
6. generate_single_test_json(expr,rule_name) 传入dict表达式和规则名，输出可用于测试的json文件
""" 

# 保存不存在的关键词
writer = open(r'.\test.csv','a+',newline='',encoding='utf-8')
csv_writer = csv.writer(writer)
none_list =[] 

# 加载json
word_json = {
    "Regex":{},
    "Keyword":{},
    "Dict":{}
}

fileExt = {
    
}

def load_words_json(rule_excel_path,sheet_num):
    workbook = openpyxl.load_workbook(rule_excel_path)
    worksheet = workbook.worksheets[sheet_num]
    for row in worksheet.iter_rows(min_row=3,max_col=9 ,values_only=True):
        # print(row)
        if row[0] == None:
            continue
        word_type = row[0]
        if word_type == None:
            continue
        id = int(row[1])
        name = row[2]
        label = row[3]
        reg = row[5]
        script = row[6]
        count = row[7]
        if row[8] == 'true':
            repeat = True
        else:
            repeat = False
        # 指定类型文件涉及的文件后缀
        if id<500:
            if label not in fileExt:
                fileExt[label] = [name]
            else:
                tmp = fileExt[label]
                tmp.append(name)
                fileExt[label] = tmp 
                
        if word_type == 'Regex':
            # 将正则转为base64编码
            if reg != None:
                reg_bytes = reg.encode('utf-8')
                encoded_bytes = base64.b64encode(reg_bytes)
                reg_encode = encoded_bytes.decode('utf-8')
            else:
                reg_encode = ''
            # 将lua脚本转成base64编码
            if script != None:
                script_bytes = script.encode('utf-8')
                encoded_bytes = base64.b64encode(script_bytes)
                script_encode = encoded_bytes.decode('utf-8')
            else:
                script_encode = ''
            json_data = {
                "LibraryId":str(id),
                "LibraryType":"Regex",
                "LibraryName":name,
                "Hits":count,
                "Repeat":repeat,
                "RegexItem":{
                    "Regex":reg_encode,
                    "Lua":script_encode
                }
            }
        elif word_type == 'Keyword':
            json_data = {
                "LibraryId":str(id),
                "LibraryType":"Keyword",
                "LibraryName":name,
                "Hits":count,
                "Repeat":repeat,
                "KeywordList":reg.split(',')
            }
        elif word_type == 'Dict':
            reg = json.loads(reg)
            dict_list = []
            for sub_dict in reg:
                dict_list.append({"Keyword":sub_dict,"Weight":reg[sub_dict]})
            json_data = {
                "LibraryId":str(id),
                "LibraryType":"Dict",
                "LibraryName":name,
                "Hits":count,
                "Repeat":repeat,
                "DictList":dict_list
            }
        word_json[word_type][name]=json_data

# 将字典表达式转为字符串类型
def generate_mini_expr(doc_part,exp,kw_list,reg_list,dict_list):
    # print("now parsing:",exp)
    and_or = exp[0] # and or count
    type = exp[1] # kw reg dict 
    count = exp[2] 
    sub_exp_format = ""
    final_exp = ""
    if type == 'kw':
        sub_exp_format = "KW({},\"{}\")"
    elif type == 'reg':
        sub_exp_format = "REGEX({},\"{}\")"
    elif type == 'dict':
        sub_exp_format = "DICT({},\"{}\")"
    sensitive_id = []
    for d in exp[3:]:
        if type == 'kw':
            if d not in word_json['Keyword']:
                if ('Keyword',d) not in none_list:
                    none_list.append(('Keyword',d))
                    csv_writer.writerow(['Keyword',d])
            else:
                sensitive_id.append(word_json['Keyword'][d]['LibraryId'])
                if word_json['Keyword'][d] not in kw_list:
                    kw_list.append(word_json['Keyword'][d])
        if type == 'reg':
            if d not in word_json['Regex']:
                if ('Regex',d) not in none_list:
                    none_list.append(('Regex',d))
                    csv_writer.writerow(['Regex',d])
                    print('reg',d)
            else:
                sensitive_id.append(word_json['Regex'][d]['LibraryId'])
                if word_json["Regex"][d] not in reg_list:
                    reg_list.append(word_json["Regex"][d])
        if type == 'dict':
            if d not in word_json['Dict']:
                if ('Dict',d) not in none_list:
                    none_list.append(('Dict',d))
                    csv_writer.writerow(['Dict',d])
            else:
                sensitive_id.append(word_json['Dict'][d]['LibraryId'])
                if word_json['Dict'][d] not in dict_list:
                    dict_list.append(word_json['Dict'][d])
        # sensitive_id.append(d)
    res = ",".join(sub_exp_format.format(doc_part,d) for d in sensitive_id)
    # print(res)
    if and_or == "and":
        final_exp = "all([%s],{#})" % res #.format(res)
    elif and_or == "or":
        final_exp = "any([%s],{#})" % res #.format(res)
    elif and_or == "count":
        final_exp = "count([%s],{#})>=%s" % (res,count) #.format(res,count)
    # print(final_exp)
    return final_exp,kw_list,reg_list,dict_list

# 输出：该dict转为str类型的表达式
def generate_complete_expr(expr):
    # 文件名表达式
    kw_list = []
    reg_list = []
    dict_list = []
    final_exp_str = []
    # 分类转换字符串
    # 文件内容
    # doc.Name 和 doc.Content
    doc_content = ['doc.Name','doc.Content']
    doc_content_list = []
    for content in doc_content:
        sub_exp = []
        if content in expr:
            expr_info = expr[content]
            if len(expr_info)<2:
                continue
            expr_info_conn = expr_info[0]
            for exp in expr_info[1:]:
                res,kw_list,reg_list,dict_list = generate_mini_expr(content,exp,kw_list,reg_list,dict_list)
                sub_exp.append(res)
            if expr_info_conn == 'and':
                sub_exp_str = ' and '.join(sub_exp)
            elif expr_info_conn == 'or':
                sub_exp_str = ' or '.join(sub_exp)
            doc_content_list.append(sub_exp_str)
    if len(doc_content_list) > 0:
        final_exp_str.append('({})'.format(' and '.join(doc_content_list)))
    
    # 文件类型
    # doc.Type，需要先将大类型转换成对应的小类别
    sub_exp = []
    if 'doc.Type' in expr:
        expr_info = expr['doc.Type']
        expr_info_conn = expr_info[0]
        for exp in expr_info[3:]:
            tmp_exp = [exp[0],exp[1],exp[2]]
            for top_exp in exp[3:]:
                if top_exp not in fileExt:
                    print('该后缀未覆盖',exp)
                    tmp_exp = tmp_exp.append(top_exp)
                    continue
                fileExt_name = fileExt[top_exp]
                tmp_exp = tmp_exp+fileExt_name
            res,kw_list,reg_list,dict_list = generate_mini_expr('doc.Type',tuple(tmp_exp),kw_list,reg_list,dict_list)
            sub_exp.append(res)
        if expr_info_conn == 'and':
            sub_exp_str = ' and '.join(sub_exp)
        elif expr_info_conn == 'or':
            sub_exp_str = ' or '.join(sub_exp)  
        if len(sub_exp)>=2:
            sub_exp_str = "({})".format(sub_exp_str)
        sub_exp_str = "(doc.Encrypt == {} and doc.HideExt == {} and {})".format(expr_info[1],expr_info[2],sub_exp_str)
        final_exp_str.append(sub_exp_str)
    
    # 文件大小
    if 'doc.Size' in expr:
        size_info = expr['doc.Size']
        size_str = ' and '.join('doc.Size {}'.format(size) for size in size_info)
        final_exp_str.append('({})'.format(size_str))
        
            
    # 文件md5 
    if 'doc.Md5' in expr:
        print(expr['doc.Md5'])
        md5_list = ','.join('\"{}\"'.format(md5) for md5 in expr['doc.Md5'])
        final_exp_str.append('doc.Md5 in [{}]'.format(md5_list))
    return ' and '.join(final_exp_str),kw_list,reg_list,dict_list    

# 根据excel生成存入db的csv
# 保存敏感词
def trans_sensitive_word_csv(rule_excel_path,sheet_num):
    writer = open(r'D:\UGit\dlp-rule-data\DLP规则测试\db数据\sensitive_word.csv','w',newline='',encoding='utf-8')
    csv_writer = csv.writer(writer) #quoting=csv.QUOTE_NONE, escapechar='\\'
    # csv_writer.writerow(['id','name','rule_type','description','labels','content','hit_count','repeated'])
    workbook = openpyxl.load_workbook(rule_excel_path)
    worksheet = workbook.worksheets[sheet_num]
    for row in worksheet.iter_rows(min_row=3,max_col=10,values_only=True):
        if row[1] == None:
            continue
        id = str(int(row[1]))
        name = row[2]
        rule_type = row[0]
        description = row[4]
        labels = str(row[3].split(',')).replace('[','{').replace(']','}').replace('\'','"')
        reg = row[5]
        lua = row[6]
        # print(reg)
        if rule_type == 'Keyword':
            content = {"Keywords":reg.split(',')}
            # content = ""
            # "\{\"Keywords\":{}\}".format(reg.split(","))
        elif rule_type == 'Regex':
            if reg != None:
                reg_bytes = reg.encode('utf-8')
                encoded_bytes = base64.b64encode(reg_bytes)
                reg_encode = encoded_bytes.decode('utf-8')
            else:
                reg_encode = ''
            if lua != None:
                lua_bytes = lua.encode('utf-8')
                encoded_bytes = base64.b64encode(lua_bytes)
                lua_encode = encoded_bytes.decode('utf-8')
            else:
                lua_encode = ''
            content = {"Regex":reg_encode,"Lua":lua_encode}
        elif rule_type == 'Dict':
            dict_json_data = json.loads(reg)
            content_json_format = {"Keyword":"","Weight":10}
            dict_list = []
            for dict_data in dict_json_data:
                content_json_format["Keyword"] = dict_data
                content_json_format["Weight"] = dict_json_data[dict_data]
                dict_list.append(content_json_format)
            content = {"Dict":dict_list}
        content = json.dumps(content,ensure_ascii=False)
        csv_writer.writerow([id,name,rule_type,description,labels,str(content),str(row[7]),row[8]])
    writer.close()
    
# 保存分级分类规则
def trans_data_rule_csv(rule_excel_path,sheet_num):
    category_dict = {
        "产品设计资料":"1",
        "公司战略资料":"2",
        "人力资源文件":"3",
        "商务运营资料":"4",
        "研发代码数据":"5",
        "用户敏感信息":"6",
        "通用特殊文件":"7"
    }
    writer = open(r'D:\UGit\dlp-rule-data\DLP规则测试\db数据\data_rule.csv','w',newline='',encoding='utf-8')
    csv_writer = csv.writer(writer)
    # csv_writer.writerow(['category_id','rule_id','name','description','data_level','expression','lib_id'])
    workbook = openpyxl.load_workbook(rule_excel_path)
    worksheet = workbook.worksheets[sheet_num]
    for row in worksheet.iter_rows(min_row=2,max_col=7,values_only=True):
        if row[2] == None or row[2] == 'None':
            continue
        category_id = category_dict[row[1]]
        id = str(int(row[2]))
        name = row[3]
        description = row[4]
        data_level = row[5]
        expression,kw_list,reg_list,dict_list = generate_complete_expr(eval(row[6]))
        kw_ids = [data["LibraryId"] for data in kw_list]
        reg_ids = [data["LibraryId"] for data in reg_list]
        dict_ids = [data["LibraryId"] for data in dict_list]
        lib_id = kw_ids+reg_ids+dict_ids
        expression.replace("\"\"","\"")
        csv_writer.writerow([category_id,name,description,data_level,str(expression),str(lib_id).replace('[','{').replace(']','}').replace('\'','"')])  

# 最后确定的测试json格式
def generate_new_test_json(rule_excel_path,sheet_num):
    # writer = open(r'.\test_expr.csv','a+',newline='')
    # csv_writer=csv.writer(writer)
    # 规则json格式
    rule_json = {
        "DocInfo":{
            "Type": ".docx",
            "Size": 12345,
            "Name": "test.docx",
            "Path": "C:\\Users\\xxxxx\\Desktop\\test.docx",
            "Content":"6226984873649982"
        },
        "LibraryItems":[
            {
                "LibraryId":"",
                "LibraryType":"keyword",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "KeywordList":[]
            },
            {
                "LibraryId":"",
                "LibraryType":"keyword",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "RegexItem":{
                    "Regex":"",
                    "Lua":""
                }
            }
        ],
        "RuleItems":[
            {
                "RuleId":"",
                "RuleName":"",
                "Level":"",
                "Expr":"",
                "LibraryList":[]
            }
        ]
    }
    # 读取分级分类规则
    workbook = openpyxl.load_workbook(rule_excel_path)
    worksheet = workbook.worksheets[sheet_num]
    # 生成json
    for row in worksheet.iter_rows(min_row=2,max_col=12,values_only=True):
        if row[1] == None:
            continue
        rule_name = row[3]
        rule_id = row[2]
        level = row[5]
        expr = eval(row[6])
        str_expr,kw_list,reg_list,dict_list = generate_complete_expr(expr)
        id_list = []
        rule_json["LibraryItems"] = kw_list+reg_list+dict_list
        for word in kw_list+reg_list:
            id = word['LibraryId'] 
            id_list.append(id)
        content = str(row[9])
        content_bytes = content.encode('utf-8')
        encoded_bytes = base64.b64encode(content_bytes)
        content_encode = encoded_bytes.decode('utf-8')
        doc_type = row[10]
        if row[11] == None or row[11] == "None":
            size = 1024
        else:
            size = row[11]
        # csv_writer.writerow([str(str_expr)])
    # 保存至json文件
        rule_json["RuleItems"][0]["Expr"] = str_expr
        rule_json["RuleItems"][0]['RuleName'] = rule_name
        rule_json["RuleItems"][0]['Level'] = level
        rule_json["RuleItems"][0]["RuleId"] = str(rule_id)
        rule_json["RuleItems"][0]["LibraryList"] = id_list
        rule_json['DocInfo']['Type'] = '.'+doc_type
        rule_json['DocInfo']['Size'] = int(size)
        rule_json['DocInfo']['Content'] = content_encode
        rule_json['DocInfo']['Name'] = row[8]
        rule_json['DocInfo']['Path'] = 'C:\\Users\\xxxx\\Desktop\\'+row[8]
        with open(os.path.join(r'D:\UGit\dlp-rule-data\DLP规则测试\测试json\测试分级分类规则',rule_name+'.json'),'w',encoding='utf-8') as f:
            json.dump(rule_json,f,ensure_ascii=False,indent=4)

# 生成测试敏感词的json数据
def generate_test_word_json(rule_excel_path,sheet_num):
    # writer = open(r'.\test_expr.csv','a+',newline='')
    # csv_writer=csv.writer(writer)
    # 规则json格式
    rule_json = {
        "DocInfo":{
            "Type": ".docx",
            "Size": 12345,
            "Name": "test.docx",
            "Path": "C:\\Users\\xxxxx\\Desktop\\test.docx",
            "Content":"6226984873649982"
        },
        "LibraryItems":[
            {
                "LibraryId":"",
                "LibraryType":"keyword",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "KeywordList":[]
            },
            {
                "LibraryId":"",
                "LibraryType":"regex",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "RegexItem":{
                    "Regex":"",
                    "Lua":""
                }
            }
        ],
        "RuleItems":[
            {
                "RuleId":"",
                "RuleName":"",
                "Level":"",
                "Expr":"",
                "LibraryList":[]
            }
        ]
    }
    # 读取分级分类规则
    workbook = openpyxl.load_workbook(rule_excel_path)
    worksheet = workbook.worksheets[sheet_num]
    # 生成json
    for row in worksheet.iter_rows(min_row=3,max_col=12,values_only=True):
        if row[0] == None or row[0]!='Regex':
            continue
        # 生成规则中敏感词库部分
        rule_id = int(row[1])
        if rule_id < 500:
            continue
        rule_name = row[2]
        content_encode = ''
        script_encode = ''
        content = str(row[5])
        if content != None:
            content_bytes = content.encode('utf-8')
            encoded_bytes = base64.b64encode(content_bytes)
            content_encode = encoded_bytes.decode('utf-8')
        
        script = str(row[6])
        if script != None and script != 'None':
            script_bytes = script.encode('utf-8')
            encoded_bytes = base64.b64encode(script_bytes)
            script_encode = encoded_bytes.decode('utf-8')
        item = {
            "LibraryId":str(rule_id),
            "LibraryType":"regex",
            "LibraryName":rule_name,
            "Hits":1,
            "Repeat":True,
            "RegexItem":{
                "Regex":content_encode,
                "Lua":script_encode
            }
        }
        rule_json['LibraryItems']=[item]
        
        # 生成规则内容
        
        
        expr = "REGEX(doc.Content, \"{}\")".format(str(rule_id))
        rule_json['RuleItems'][0]['Expr'] = expr
        rule_json["RuleItems"][0]['Level'] = 'S1'
        rule_json["RuleItems"][0]['RuleName'] = row[2]
        rule_json["RuleItems"][0]['RuleId'] = str(rule_id)
        rule_json["RuleItems"][0]["LibraryList"] = [str(rule_id)]
        # 生成规则测试部分
        content = str(row[9])
        content_bytes = content.encode('utf-8')
        encoded_bytes = base64.b64encode(content_bytes)
        content_encode = encoded_bytes.decode('utf-8')
        rule_json['DocInfo']['Content'] = content_encode
        rule_name = rule_name.replace(' ','')
        with open(os.path.join(r'D:\UGit\dlp-rule-data\DLP规则测试\测试json\测试敏感词',rule_name+'.json'),'w',encoding='utf-8') as f:
            json.dump(rule_json,f,ensure_ascii=False,indent=4)

# 根据单独表达式生成json文件
def generate_single_test_json(expr,rule_name):
    # 规则json格式
    rule_json = {
        "DocInfo":{
            "Type": ".docx",
            "Size": 12345,
            "Name": "test.docx",
            "Path": "C:\\Users\\xxxxx\\Desktop\\test.docx",
            "Content":"6226984873649982"
        },
        "LibraryItems":[
            {
                "LibraryId":"",
                "LibraryType":"keyword",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "KeywordList":[]
            },
            {
                "LibraryId":"",
                "LibraryType":"keyword",
                "LibraryName":"",
                "Hits":1,
                "Repeat":True,
                "RegexItem":{
                    "Regex":"",
                    "Lua":""
                }
            }
        ],
        "RuleItems":[
            {
                "RuleId":"",
                "RuleName":"",
                "Level":"",
                "Expr":"",
                "LibraryList":[]
            }
        ]
    }
    res = generate_complete_expr(expr)
    
    # 敏感词库
    id_list = []
    rule_json["LibraryItems"] = res[1]+res[2]+res[3]
    for word in res[1]+res[2]+res[3]:
        id = word['LibraryId'] 
        id_list.append(id)
        
    # 规则内容
    rule_json["RuleItems"][0]["Expr"] = res[0]
    rule_json["RuleItems"][0]['RuleName'] = rule_name
    rule_json["RuleItems"][0]['Level'] = 'S3'
    rule_json["RuleItems"][0]["RuleId"] = '100'
    rule_json["RuleItems"][0]["LibraryList"] = id_list
    with open(os.path.join(r'C:\Users\yangqilv\Desktop\DLP测试文本\测试json\测试分级分类规则',rule_name+'.json'),'w',encoding='utf-8') as f:
        json.dump(rule_json,f,ensure_ascii=False,indent=4)

      
if __name__ == "__main__":
    rule_path = r'D:\UGit\dlp-rule-data\DLP规则测试\测试脚本\DLP文本分级分类规则.xlsx'
    load_words_json(rule_path,1)
    print(fileExt)
    # generate_test_json_file()
    # generate_all_init_exp()
    
    '''
    测试生成单个表达式
    '''
    # expr = {
    #     "doc.Name":["and",("and","kw",1,"姓名","性别","年龄","体检日期")],
    #     "doc.Content":['or',("and","kw",1,"姓名","性别","年龄","体检日期"),("or","reg",1,"大量手机号码(中国大陆)","大量手机号码(中国台湾)","大量手机号码(中国香港)"),("count","kw",2,"姓名","性别","年龄","体检日期")],
    #     "doc.Size":['>1024'],
    #     "doc.Type":['or','false','false',("or",'reg',1,"办公文档","文本文档")],
    #     "doc.Md5":['123415262','326666','33333']
    # }
    # res = generate_complete_expr(expr)
    # print(res[0])
    # print(res[1])
    # print(res[2])
    # generate_single_test_json(expr,'test')
    
    '''
    测试根据excel批量生成用于测试的json文件
    '''
    # generate_new_test_json(rule_path,2)
    # generate_test_word_json(rule_path,1)
    
    '''
    根据excel批量生成需要导入db的csv格式数据
    '''
    trans_sensitive_word_csv(rule_path,1)
    trans_data_rule_csv(rule_path,2)
    
    